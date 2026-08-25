"""
Bulk product import from an Excel (.xlsx) or CSV file.

Used by the "Import products" button on the Product changelist (see admin.py).

The column names mirror the fields in product-data/products.json, so a sheet
exported from the catalogue can be fed straight back in:

    sku, name, brand, category, subcategory, collection, highlight,
    description, tags, amazon_link, gst_pct, hero, gallery, video, features

`sku` is the identity column: a row whose sku already exists UPDATES that
product in place, a new sku CREATES one. Nothing is ever duplicated, and
re-uploading the same file is a no-op the second time (every row comes back
as "unchanged"). A bad row is reported and skipped — it never aborts the run.

Brands and categories are looked up by name or slug and are NEVER created:
a row naming an unknown brand/category is reported as an error instead, so a
typo can't litter the catalogue with junk brands.
"""
import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.utils.text import slugify

from .models import (
    Brand, Category, Marketplace, Product, ProductImage, ProductMarketplaceLink,
)

# Order matters — this is also the header row of the downloadable template.
COLUMNS = [
    'sku', 'name', 'brand', 'category', 'subcategory', 'collection', 'highlight',
    'description', 'tags', 'amazon_link', 'gst_pct', 'hero', 'gallery', 'video',
    'features',
]

# Two real rows from the catalogue, so a non-technical user can see the shape
# of every column (multi-value cells, the "Title | detail" feature syntax, and
# paths rather than uploaded files for hero/gallery/video).
TEMPLATE_EXAMPLE_ROWS = [
    [
        'LI001A',
        'CRYSTAL PLASTIC ABS LIGHTER, MULTICOLOUR',
        'crystal',
        'kitchenware',
        'lighters',
        'LIGHTERS',
        'CRYSTAL ABS LIGHTER (PLASTIC BODY)',
        'A Durable Plastic Abs Lighter, Ideal For Gas Stoves And Kitchen Use.',
        'Kitchen Lighter, Abs Lighter, Gas Stove Lighter',
        'https://www.amazon.in/dp/B00J4YDI52',
        '0.18',
        'product-photos/LI001A/hero.webp',
        'product-photos/LI001A/g1.webp, product-photos/LI001A/g2.webp, product-photos/LI001A/g3.webp',
        'product-photos/LI001A/video.mp4',
        'Sturdy ABS Body | Impact-resistant plastic that survives daily kitchen use\n'
        'Reliable Ignition | Consistent spark every time you press\n'
        'Easy Grip | Ergonomic shape for a comfortable hold',
    ],
    [
        'LI002A',
        'CRYSTAL ARISTO STAINLESS STEEL LIGHTER, MULTICOLOUR',
        'crystal',
        'kitchenware',
        'lighters',
        'LIGHTERS',
        'CRYSTAL SS LIGHTER (ARISTO)',
        'The Crystal Aristo Stainless Steel Lighter combines elegance and functionality '
        'in one sleek design, with a reliable ignition system for gas stoves, grills and candles.',
        'Stainless Steel Gas Lighter | Crystal Aristo Lighter | Durable Gas Stove Lighter',
        'https://www.amazon.in/dp/B00INQOQ3U',
        '0.18',
        'product-photos/LI002A/hero.webp',
        'product-photos/LI002A/g1.webp | product-photos/LI002A/g2.webp',
        'product-photos/LI002A/video.mp4',
        'Sturdy Stainless Steel Construction | Built to last in a busy kitchen\n'
        'Refillable and Reusable | Top up the gas instead of throwing it away\n'
        'Multicolour Finish | A modern touch on the counter',
    ],
]

DEFAULT_FEATURE_ICON = 'shield'   # Product.html falls back to this icon anyway
_LIST_SPLIT = re.compile(r'[,|\n]+')


class ImportError_(Exception):
    """A row-level problem: reported against the row, never fatal."""


# ── cell helpers ───────────────────────────────────────────────────────────

def _clean(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _split_list(value):
    """tags / gallery: comma-separated or pipe-separated (newlines too)."""
    return [part.strip() for part in _LIST_SPLIT.split(value) if part.strip()]


def _parse_features(value):
    """
    "Title | detail" per line, or a plain comma-separated list of titles.
    Stored as Product.features == [[icon, title, detail], ...].
    """
    if not value:
        return []
    lines = [ln.strip() for ln in value.splitlines() if ln.strip()]
    if len(lines) == 1 and '|' not in lines[0]:
        lines = [part.strip() for part in lines[0].split(',') if part.strip()]
    features = []
    for line in lines:
        parts = [part.strip() for part in line.split('|')]
        if len(parts) >= 3:
            # same "icon | title | detail" form the product edit screen uses
            icon, title, detail = parts[0] or DEFAULT_FEATURE_ICON, parts[1], parts[2]
        else:
            icon, title = DEFAULT_FEATURE_ICON, parts[0]
            detail = parts[1] if len(parts) > 1 else ''
        if title:
            features.append([icon, title, detail])
    return features


def _parse_gst(value):
    """Accepts 0.18, 18, or "18%" — all mean 18%."""
    if not value:
        return None
    text = value.replace('%', '').strip()
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ImportError_(f"gst_pct '{value}' is not a number")
    if number > 1:
        number = number / Decimal('100')
    if number < 0 or number > 1:
        raise ImportError_(f"gst_pct '{value}' is out of range")
    return number.quantize(Decimal('0.0001'))


# ── file parsing ───────────────────────────────────────────────────────────

def _normalise_header(cell):
    return re.sub(r'[^a-z0-9]+', '_', _clean(cell).lower()).strip('_')


def read_rows(uploaded_file):
    """
    Return [{column: value}, ...] for a .xlsx or .csv upload.
    Unknown columns are ignored; the header is matched case-insensitively.
    Raises ValueError for a file we can't read at all.
    """
    name = (uploaded_file.name or '').lower()
    data = uploaded_file.read()

    if name.endswith('.xlsx'):
        table = _read_xlsx(data)
    elif name.endswith('.csv'):
        table = _read_csv(data)
    else:
        raise ValueError('Please upload a .xlsx or .csv file.')

    if not table:
        raise ValueError('The file is empty.')

    header = [_normalise_header(cell) for cell in table[0]]
    known = set(COLUMNS)
    if not known & set(header):
        raise ValueError(
            'No recognised columns found in the header row. Expected at least: '
            + ', '.join(COLUMNS[:4]) + ' …'
        )

    rows = []
    for raw in table[1:]:
        row = {}
        for index, key in enumerate(header):
            if key in known:
                row[key] = _clean(raw[index]) if index < len(raw) else ''
        if any(row.values()):
            rows.append(row)
    return rows


def _read_xlsx(data):
    import openpyxl
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [list(r) for r in sheet.iter_rows(values_only=True)]


def _read_csv(data):
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('Could not decode the CSV file.')
    return [row for row in csv.reader(io.StringIO(text, newline=''))]


# ── lookups (never create) ─────────────────────────────────────────────────

def _find_brand(value):
    if not value:
        raise ImportError_('brand is required')
    brand = (Brand.objects.filter(slug__iexact=value).first()
             or Brand.objects.filter(name__iexact=value).first()
             or Brand.objects.filter(slug__iexact=slugify(value)).first())
    if brand is None:
        raise ImportError_(f"unknown brand '{value}' — add it under Brands first")
    return brand


def _match_category(value):
    return (Category.objects.filter(slug__iexact=value).first()
            or Category.objects.filter(name__iexact=value).first()
            or Category.objects.filter(slug__iexact=slugify(value)).first())


def _find_category(category_value, subcategory_value):
    """subcategory wins when given; the product hangs off the deepest match."""
    if subcategory_value:
        parent = _match_category(category_value) if category_value else None
        candidates = Category.objects.filter(parent=parent) if parent else Category.objects.all()
        wanted = {subcategory_value.lower(), slugify(subcategory_value)}
        for candidate in candidates:
            if candidate.slug.lower() in wanted or candidate.name.lower() in wanted:
                return candidate
        combined = _match_category(f'{category_value}-{subcategory_value}') if category_value else None
        if combined is not None:
            return combined
        raise ImportError_(
            f"unknown subcategory '{subcategory_value}' — add it under Categories first")
    if not category_value:
        raise ImportError_('category is required')
    category = _match_category(category_value)
    if category is None:
        raise ImportError_(f"unknown category '{category_value}' — add it under Categories first")
    return category


# ── the import itself ──────────────────────────────────────────────────────

def _apply_gallery(product, paths):
    """
    hero/gallery hold a path or URL, not an uploaded file, so the paths are
    stored as the ProductImage file names (nothing is downloaded) — the same
    convention the rest of the catalogue already uses, hero first and flagged
    is_hero. Returns True when the gallery actually changed.
    """
    existing = list(product.images.filter(variant__isnull=True).order_by('order', 'id'))
    if [image.image.name for image in existing] == paths:
        return False
    for image in existing:
        image.delete()
    for order, path in enumerate(paths):
        ProductImage.objects.create(
            product=product, image=path, is_hero=(order == 0), order=order)
    return True


def _link_amazon(product, url):
    """Keep the Amazon marketplace link in step — but only if that marketplace
    already exists; we don't invent one."""
    if not url:
        return False
    marketplace = Marketplace.objects.filter(slug='amazon').first()
    if marketplace is None:
        return False
    link = ProductMarketplaceLink.objects.filter(
        product=product, marketplace=marketplace).first()
    if link is None:
        ProductMarketplaceLink.objects.create(
            product=product, marketplace=marketplace, url=url)
        return True
    if link.url != url:
        link.url = url
        link.save(update_fields=['url'])
        return True
    return False


def _row_values(row):
    """Map a cleaned row onto Product field values. Only columns actually
    present (and non-blank) are returned, so an update never blanks a field
    the sheet didn't mention."""
    values = {}
    if row.get('name'):
        values['name'] = row['name'][:200]
    if row.get('brand') is not None and row.get('brand') != '':
        values['brand'] = _find_brand(row['brand'])
    if row.get('category') or row.get('subcategory'):
        values['category'] = _find_category(row.get('category', ''), row.get('subcategory', ''))
    if row.get('collection'):
        values['collection_name'] = row['collection'][:200]
    if row.get('highlight'):
        values['highlight'] = row['highlight'][:300]
        values['short_description'] = row['highlight'][:300]
    if row.get('description'):
        values['overview'] = row['description']
    if row.get('tags'):
        values['tags'] = _split_list(row['tags'])
    if row.get('amazon_link'):
        values['amazon_link'] = row['amazon_link'][:500]
    if row.get('gst_pct'):
        values['gst_pct'] = _parse_gst(row['gst_pct'])
    if row.get('hero'):
        values['image_url'] = row['hero'][:500]
    if row.get('video'):
        values['video_url'] = row['video'][:500]
    if row.get('features'):
        values['features'] = _parse_features(row['features'])
    return values


def import_rows(rows):
    """
    Apply parsed rows. Returns (results, summary) where results is a list of
    dicts: {line, sku, name, status, detail}. status is one of
    created / updated / unchanged / skipped.
    """
    results = []
    summary = {'created': 0, 'updated': 0, 'unchanged': 0, 'skipped': 0}

    for index, row in enumerate(rows, start=2):   # row 1 is the header
        sku = row.get('sku', '')
        record = {'line': index, 'sku': sku, 'name': row.get('name', ''),
                  'status': 'skipped', 'detail': ''}
        if not sku:
            record['detail'] = 'no sku — the sku column identifies the product'
            summary['skipped'] += 1
            results.append(record)
            continue

        product = Product.objects.filter(sku__iexact=sku).first()
        try:
            values = _row_values(row)
            if product is None and not values.get('name'):
                raise ImportError_('new product needs a name')
            if product is None and 'brand' not in values:
                raise ImportError_('new product needs a brand')
            if product is None and 'category' not in values:
                raise ImportError_('new product needs a category')
        except ImportError_ as exc:
            record['detail'] = str(exc)
            summary['skipped'] += 1
            results.append(record)
            continue

        # The catalogue stores the hero as the first (is_hero) image row, so the
        # hero path leads the list when both columns are filled in.
        gallery = None
        if row.get('gallery'):
            gallery = _split_list(row['gallery'])
            hero = row.get('hero', '').strip()
            if hero and hero not in gallery:
                gallery.insert(0, hero)

        if product is None:
            product = Product(sku=sku, slug=slugify(sku), **values)
            product.save()
            if gallery:
                _apply_gallery(product, gallery)
            _link_amazon(product, values.get('amazon_link', ''))
            record.update(status='created', name=product.name, detail='new product')
            summary['created'] += 1
            results.append(record)
            continue

        changed = []
        for field, value in values.items():
            if getattr(product, field) != value:
                setattr(product, field, value)
                changed.append(field)
        if changed:
            product.save()
        if gallery is not None and _apply_gallery(product, gallery):
            changed.append('gallery')
        if _link_amazon(product, values.get('amazon_link', '')):
            changed.append('amazon marketplace link')

        record['name'] = product.name
        if changed:
            record.update(status='updated', detail='changed: ' + ', '.join(changed))
            summary['updated'] += 1
        else:
            record.update(status='unchanged', detail='already up to date')
            summary['unchanged'] += 1
        results.append(record)

    return results, summary


# ── the downloadable template ──────────────────────────────────────────────

def template_csv_bytes():
    buffer = io.StringIO(newline='')
    writer = csv.writer(buffer)
    writer.writerow(COLUMNS)
    for row in TEMPLATE_EXAMPLE_ROWS:
        writer.writerow(row)
    return buffer.getvalue().encode('utf-8-sig')


def template_xlsx_bytes():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Products'
    sheet.append(COLUMNS)
    for row in TEMPLATE_EXAMPLE_ROWS:
        sheet.append(row)

    header_fill = PatternFill('solid', fgColor='ED3338')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center')
    widths = {'sku': 12, 'name': 46, 'brand': 12, 'category': 16, 'subcategory': 16,
              'collection': 14, 'highlight': 34, 'description': 52, 'tags': 34,
              'amazon_link': 34, 'gst_pct': 9, 'hero': 30, 'gallery': 46,
              'video': 30, 'features': 60}
    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = widths[column]
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    sheet.freeze_panes = 'A2'

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
