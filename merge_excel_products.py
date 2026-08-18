import json
import re
import openpyxl

XLSX_PATH = r"C:\Users\prasa\Downloads\Crystal Product List for Techinfinity Final (1).xlsx"
JSON_PATH = r"C:\Website\crystal\product-data\products.json"

SHEET_CONFIG = {
    "Lighter": {"category": "kitchenware", "subcategory": "lighters"},
    "Knife": {"category": "kitchenware", "subcategory": "knives"},
    "Peeler": {"category": "kitchenware", "subcategory": "peelers"},
    "Chopping Boards": {"category": "kitchenware", "subcategory": "chopping-boards"},
    "Trolley": {"category": "kitchenware", "subcategory": "trolleys"},
    "Manual Kitchen Appliances": {"category": "kitchenware", "subcategory": "manual-appliances"},
    "Kitchen Tools": {"category": "kitchenware", "subcategory": "kitchen-tools"},
    "Cutlery": {"category": "kitchenware", "subcategory": "cutlery"},
    "Servers": {"category": "kitchenware", "subcategory": "servers"},
    "Water Filter": {"category": "kitchenware", "subcategory": "water-filter"},
    "Water Bottle": {"category": "water-bottle", "subcategory": None},
    "Oil Pourer & Sprayer": {"category": "oil-pourer", "subcategory": None},
    "Wood Range": {"category": "wood-range", "subcategory": None},
    "Cookware": {"category": "cookware", "subcategory": "__from_material__"},
    "Pressure Cooker": {"category": "pressure-cooker", "subcategory": None},
    "Electric Appliances": {"category": "electric-appliances", "subcategory": "__from_type__"},
    "Cooktop": {"category": "cooktop", "subcategory": None},
    "LUNCHBOX": {"category": "lunch-box", "subcategory": None},
    "cleaningaid": {"category": "cleaning-aid", "subcategory": "__from_subcat_col__"},
}

MATERIAL_TO_SUB = {
    "Triply": "tripro", "Sandwich Bottom Steel": "sandwich-bottom-steel",
    "CAST IRON": "cast-iron", "Nonstick - Mini": "non-stick-mini",
    "Nonstick": "non-stick", "Hard Anodised": "hard-anodised",
}
TYPE_TO_SUB_EA = {
    "Rice Cooker": "rice-cooker", "OTG": "otg", "Food Processor": "food-processor",
    "Kettle": "kettle", "JMG": "jmg", "Chimney": "chimney", "Iron": "iron",
    "Air Fryer": "air-fryer", "Ice Cream Maker": "ice-cream-maker",
}
CA_SUBCAT_TO_SLUG = {
    "Hand Held Mop": "hand-held-mops", "Plunger": "plunger", "Brush": "brush",
    "Scrubber & Scourer": "scrubber", "Spin Mop": "spin-mops", "Wipe": "wipe",
    "Sink Organiser": "sink-organiser", "Broom & Dustpan": "brooms",
    "Bin": "bins", "Wiper": "wipers",
}

# columns to ignore when building the generic "filters" dict
IGNORE_COLS = {
    "brand", "item category", "new product code", "product code", "item description",
    "item description with pic", "amzon y/n", "amzon \ny/n", "store", "link",
    "manufacturer name", "manufacturer address", "brand \nname", "hdn code", "hsn code",
    "product name (max 100 characters)", "unit \nquantity", "uom", "weight \nin gram",
    "country \nof origin", "length, cm \n(packaging)", "breadth, cm \n(packaging)",
    "height, cm \n(packaging)", "case pack \nsize qty", "carton, bottle, \npouch, jar, pack",
    "gst %", "mrp", "product \ndescription", "customer service no.",
    "customer service email id", "search keywords\n( all keywords seperated by a comma)",
    "item subgroup", "item subcategory", "category", "sub category",
}


def norm(s):
    return (s or "").strip().lower() if isinstance(s, str) else s


def slugify_generic(s):
    return re.sub(r"[^a-z0-9]+", "-", str(s).strip().lower()).strip("-")


def brand_slug(raw):
    r = norm(raw)
    if not r:
        return "crystal"
    if "crystalina" in r:
        return "crystalina"
    if "sparkmate" in r or "spark mate" in r:
        return "sparkmate"
    if "valmate" in r or "val mate" in r:
        return "valmate"
    return "crystal"


def extract_rows(sheet_name, ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_row_idx = 0
    for i, r in enumerate(rows[:5]):
        cells = {norm(c) for c in r if c is not None}
        if "new product code" in cells or "product code" in cells:
            header_row_idx = i
            break
    header = [h if h is not None else "" for h in rows[header_row_idx]]
    header_norm = [norm(h) for h in header]
    rows = rows[header_row_idx:]
    cfg = SHEET_CONFIG[sheet_name]
    out = []

    def col(*names):
        for n in names:
            nn = norm(n)
            if nn in header_norm:
                return header_norm.index(nn)
        return None

    code_i = col("NEW PRODUCT CODE", "PRODUCT CODE")
    desc_i = col("ITEM DESCRIPTION", "ITEM DESCRIPTION WITH PIC")
    pname_i = col("Product Name (max 100 characters)")
    brand_i = col("Brand")
    amzy_i = col("Amzon Y/N", "Amzon \nY/N")
    link_i = col("LINK")
    pdesc_i = col("Product \nDescription")
    kw_i = col("Search Keywords\n( All keywords seperated by a comma)")
    gst_i = col("GST %")
    mrp_i = col("MRP")
    mat_i = col("Material", "Material ")
    type_i = col("Type")
    subcat_i = col("Sub Category")

    if code_i is None:
        return []

    for r in rows[1:]:
        code = r[code_i] if code_i < len(r) else None
        if not code or not str(code).strip():
            continue
        sku = str(code).strip().upper()
        name = None
        if pname_i is not None and pname_i < len(r) and r[pname_i]:
            name = str(r[pname_i]).strip()
        elif desc_i is not None and desc_i < len(r) and r[desc_i]:
            name = str(r[desc_i]).strip()
        if not name:
            name = sku

        highlight = None
        if pdesc_i is not None and pdesc_i < len(r) and r[pdesc_i]:
            highlight = str(r[pdesc_i]).strip()
        elif desc_i is not None and desc_i < len(r) and r[desc_i]:
            highlight = str(r[desc_i]).strip()

        tags = []
        if kw_i is not None and kw_i < len(r) and r[kw_i]:
            tags = [t.strip() for t in str(r[kw_i]).split(",") if t.strip()]

        amazon_link = None
        if link_i is not None and link_i < len(r) and r[link_i] and str(r[link_i]).strip().startswith("http"):
            amazon_link = str(r[link_i]).strip()

        gst_pct = None
        if gst_i is not None and gst_i < len(r) and isinstance(r[gst_i], (int, float)):
            gst_pct = r[gst_i]
        mrp = None
        if mrp_i is not None and mrp_i < len(r) and isinstance(r[mrp_i], (int, float)):
            mrp = r[mrp_i]

        # subcategory resolution
        sub = cfg["subcategory"]
        if sub == "__from_material__":
            mat_val = r[mat_i] if mat_i is not None and mat_i < len(r) else None
            sub = MATERIAL_TO_SUB.get(str(mat_val).strip() if mat_val else "", None)
        elif sub == "__from_type__":
            type_val = r[type_i] if type_i is not None and type_i < len(r) else None
            sub = TYPE_TO_SUB_EA.get(str(type_val).strip() if type_val else "", None)
        elif sub == "__from_subcat_col__":
            sc_val = r[subcat_i] if subcat_i is not None and subcat_i < len(r) else None
            sub = CA_SUBCAT_TO_SLUG.get(str(sc_val).strip() if sc_val else "", None)

        # generic filters from remaining attribute-ish columns
        filters = {}
        for idx, hn in enumerate(header_norm):
            if not hn or hn in IGNORE_COLS:
                continue
            if idx >= len(r) or r[idx] is None:
                continue
            val = r[idx]
            if isinstance(val, str) and not val.strip():
                continue
            key = slugify_generic(header[idx]).replace("-", "_")
            filters[key] = val if not isinstance(val, str) else val.strip()

        out.append({
            "sku": sku,
            "name": name.upper() if name else sku,
            "brand": brand_slug(r[brand_i] if brand_i is not None and brand_i < len(r) else None),
            "category": cfg["category"],
            "subcategory": sub,
            "highlight": highlight or name,
            "description": highlight or name,
            "tags": tags,
            "gst_pct": gst_pct,
            "mrp": mrp,
            "amazon_link": amazon_link,
            "filters": filters,
        })
    return out


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    excel_products = []
    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows = extract_rows(sheet_name, wb[sheet_name])
        excel_products.extend(rows)
        print(f"{sheet_name}: {len(rows)} rows extracted")

    print(f"\nTotal extracted from Excel: {len(excel_products)}")

    with open(JSON_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    existing = raw["products"]
    existing_by_sku = {(p.get("sku") or "").strip().upper(): p for p in existing}

    added, updated = 0, 0
    for ep in excel_products:
        sku = ep["sku"]
        if sku in existing_by_sku:
            p = existing_by_sku[sku]
            # Excel is authoritative for these fields; keep existing hero/gallery/match_tier/id untouched
            p["name"] = ep["name"] or p.get("name")
            p["brand"] = ep["brand"] or p.get("brand")
            p["category"] = ep["category"] or p.get("category")
            if ep["subcategory"]:
                p["subcategory"] = ep["subcategory"]
            p["highlight"] = ep["highlight"] or p.get("highlight")
            p["description"] = ep["description"] or p.get("description")
            if ep["tags"]:
                p["tags"] = ep["tags"]
            if ep["gst_pct"] is not None:
                p["gst_pct"] = ep["gst_pct"]
            if ep["mrp"] is not None:
                p["mrp"] = ep["mrp"]
            if ep["amazon_link"]:
                p["amazon_link"] = ep["amazon_link"]
            if ep["filters"]:
                p["filters"] = {**(p.get("filters") or {}), **ep["filters"]}
            updated += 1
        else:
            new_p = {
                "sku": ep["sku"],
                "name": ep["name"],
                "brand": ep["brand"],
                "category": ep["category"],
                "subcategory": ep["subcategory"],
                "collection": (ep["subcategory"] or ep["category"] or "").upper(),
                "highlight": ep["highlight"],
                "description": ep["description"],
                "tags": ep["tags"],
                "gst_pct": ep["gst_pct"],
                "mrp": ep["mrp"],
                "amazon_link": ep["amazon_link"],
                "filters": ep["filters"],
                "hero": None,
                "gallery": [],
                "match_tier": "from_excel",
                "id": ep["sku"].lower(),
            }
            existing.append(new_p)
            existing_by_sku[sku] = new_p
            added += 1

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(raw, f, indent=2, ensure_ascii=False)

    print(f"\nAdded (new): {added}")
    print(f"Updated (existing): {updated}")
    print(f"Total products now in products.json: {len(existing)}")


if __name__ == "__main__":
    main()
