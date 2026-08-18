import json
import openpyxl

XLSX_PATH = r"C:\Users\prasa\Downloads\Crystal Product List for Techinfinity Final (1).xlsx"
JSON_PATH = r"C:\Website\crystal\product-data\products.json"

SHEET_NAMES = [
    "Lighter", "Knife", "Peeler", "Chopping Boards", "Trolley",
    "Manual Kitchen Appliances", "Kitchen Tools", "Cutlery", "Servers",
    "Water Filter", "Water Bottle", "Oil Pourer & Sprayer", "Wood Range",
    "Cookware", "Pressure Cooker", "Electric Appliances", "Cooktop", "cleaningaid",
]


def norm(s):
    return (s or "").strip().lower() if isinstance(s, str) else s


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    fixes = {}  # sku -> (short, long)

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row_idx = 0
        for i, r in enumerate(rows[:5]):
            cells = {norm(c) for c in r if c is not None}
            if "new product code" in cells or "product code" in cells:
                header_row_idx = i
                break
        header = [h if h is not None else "" for h in rows[header_row_idx]]
        header_norm = [norm(h) for h in header]

        def col(*names):
            for n in names:
                nn = norm(n)
                if nn in header_norm:
                    return header_norm.index(nn)
            return None

        code_i = col("NEW PRODUCT CODE", "PRODUCT CODE")
        short_i = col("ITEM DESCRIPTION", "ITEM DESCRIPTION WITH PIC")
        long_i = col("Product \nDescription")
        if code_i is None:
            continue

        for r in rows[header_row_idx + 1:]:
            code = r[code_i] if code_i < len(r) else None
            if not code or not str(code).strip():
                continue
            sku = str(code).strip().upper()
            short_desc = str(r[short_i]).strip() if short_i is not None and short_i < len(r) and r[short_i] else ""
            long_desc = str(r[long_i]).strip() if long_i is not None and long_i < len(r) and r[long_i] else ""
            fixes[sku] = (short_desc, long_desc)

    with open(JSON_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    items = raw["products"]

    fixed_count = 0
    for p in items:
        sku = (p.get("sku") or "").strip().upper()
        if sku not in fixes:
            continue
        short_desc, long_desc = fixes[sku]
        highlight = short_desc or long_desc or p.get("name", "")
        description = long_desc or short_desc or p.get("name", "")
        if p.get("highlight") != highlight or p.get("description") != description:
            p["highlight"] = highlight
            p["description"] = description
            fixed_count += 1

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(raw, f, indent=2, ensure_ascii=False)

    print(f"Fixed highlight/description for {fixed_count} products")


if __name__ == "__main__":
    main()
